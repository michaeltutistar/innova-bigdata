"""
Sistema Web MVP de Registro de Líderes y Sufragantes
Backend con FastAPI - API REST para gestión de datos electorales
"""

import os
import logging
import secrets
import smtplib
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from typing import List, Optional
import json
from fastapi import FastAPI, Depends, HTTPException, status, BackgroundTasks, Query, UploadFile, File, Form
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from sqlalchemy import create_engine, Column, Integer, String, Text, Boolean, DateTime, ForeignKey, func, text, case
from sqlalchemy.orm import sessionmaker, declarative_base, relationship, Session
from sqlalchemy.exc import IntegrityError
from pydantic import BaseModel, Field, field_validator
import bcrypt
from jose import JWTError, jwt
import httpx
import openpyxl
from io import BytesIO
from urllib.parse import quote_plus

# =====================
# CONFIGURACIÓN
# =====================

SECRET_KEY = os.getenv("SECRET_KEY", "tu_secret_key_super_segura_aqui_cambiar_en_produccion")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", "240"))

VERIFIK_API_URL = os.getenv("VERIFIK_API_URL", "https://api.verifik.co/v2/co/registraduria/votacion")
VERIFIK_TOKEN = os.getenv("VERIFIK_TOKEN", "")

CORS_ORIGINS = os.getenv("CORS_ORIGINS", "https://innovabigdata.com,http://localhost:5173").split(",")

# Recuperación de contraseña por email (Gmail)
GMAIL_USER = os.getenv("GMAIL_USER", "")
GMAIL_APP_PASSWORD = os.getenv("GMAIL_APP_PASSWORD", "")
RESET_PASSWORD_BASE_URL = os.getenv("RESET_PASSWORD_BASE_URL", "https://www.innovabigdata.com").rstrip("/")

# DATABASE_URL: prioridad 1) AWS Secrets Manager, 2) RDS_* del .env, 3) DATABASE_URL
def _get_database_url() -> str:
    # 1) Si está configurado el secret de RDS en Secrets Manager, obtener credenciales de ahí (rotación automática)
    secret_arn = os.getenv("AWS_RDS_SECRET_ARN")
    secret_name = os.getenv("AWS_RDS_SECRET_NAME")
    secret_id = secret_arn or secret_name
    if secret_id:
        try:
            import boto3
            import json
            client = boto3.client("secretsmanager", region_name=os.getenv("AWS_REGION", "us-east-2"))
            response = client.get_secret_value(SecretId=secret_id)
            secret = json.loads(response["SecretString"])
            user = secret.get("username", "postgres")
            password = secret.get("password", "")
            host = secret.get("host") or secret.get("hostname") or os.getenv("RDS_HOST")
            port = secret.get("port", 5432)
            dbname = secret.get("dbname") or secret.get("dbClusterIdentifier") or os.getenv("RDS_DB", "innovabigdata")
            if host and password is not None:
                safe_pwd = quote_plus(str(password))
                return f"postgresql://{user}:{safe_pwd}@{host}:{port}/{dbname}?sslmode=require"
        except Exception as e:
            logging.warning("No se pudo obtener el secret de RDS desde Secrets Manager: %s. Usando RDS_* o DATABASE_URL.", e)

    # 2) Variables RDS_* en .env
    _rds_host = os.getenv("RDS_HOST")
    if _rds_host:
        _rds_user = os.getenv("RDS_USER", "postgres")
        _rds_password = os.getenv("RDS_PASSWORD", "")
        _rds_db = os.getenv("RDS_DB", "innovabigdata")
        _rds_port = os.getenv("RDS_PORT", "5432")
        _safe_password = quote_plus(_rds_password)
        return f"postgresql://{_rds_user}:{_safe_password}@{_rds_host}:{_rds_port}/{_rds_db}?sslmode=require"

    # 3) DATABASE_URL directa
    return os.getenv("DATABASE_URL", "postgresql://postgres:postgres@db:5432/innovabigdata")

DATABASE_URL = _get_database_url()

# =====================
# LOGGING
# =====================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# =====================
# BASE DE DATOS
# =====================

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class Usuario(Base):
    __tablename__ = "usuarios"

    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    password_hash = Column(String(255), nullable=False)
    rol = Column(String(20), nullable=False)
    activo = Column(Boolean, default=True)
    fecha_creacion = Column(DateTime, default=datetime.utcnow)
    ultimo_login = Column(DateTime)
    reset_token = Column(String(64), nullable=True, index=True)
    reset_token_expires = Column(DateTime, nullable=True)
    email = Column(String(255), nullable=True)

class Lider(Base):
    __tablename__ = "lideres"

    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(Text, nullable=False)
    cedula = Column(String(10), unique=True, index=True, nullable=False)
    edad = Column(Integer, nullable=False)
    celular = Column(String(10), nullable=False)
    direccion = Column(Text, nullable=False)
    genero = Column(String(10), nullable=False)
    departamento = Column(Text, nullable=False)
    municipio = Column(Text, nullable=False)
    barrio = Column(Text)
    zona_influencia = Column(Text)
    tipo_liderazgo = Column(String(50))
    usuario_registro = Column(String(100))
    fecha_registro = Column(DateTime, default=datetime.utcnow)
    activo = Column(Boolean, default=True)

    sufragantes = relationship("Sufragante", back_populates="lider")
    incidencias_carga_masiva = relationship("CargaMasivaIncidencia", back_populates="lider")

class Sufragante(Base):
    __tablename__ = "sufragantes"

    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(Text, nullable=False)
    cedula = Column(String(10), unique=True, index=True, nullable=False)
    edad = Column(Integer, nullable=False)
    celular = Column(String(10), nullable=True)  # Opcional: "No tiene"
    direccion_residencia = Column(Text, nullable=False)
    genero = Column(String(10), nullable=True)  # NULL para registro masivo (Excel sin género)

    # Datos de Verifik
    departamento = Column(Text)
    municipio = Column(Text)
    lugar_votacion = Column(Text)
    mesa_votacion = Column(Text)
    direccion_puesto = Column(Text)

    # Estado de validación
    estado_validacion = Column(String(20), nullable=False)
    # Campos que no coincidieron con Verifik (JSON array, solo si estado=revision)
    discrepancias_verifik = Column(Text, nullable=True)

    # Referencias
    lider_id = Column(Integer, ForeignKey("lideres.id"), index=True)
    usuario_registro = Column(String(100))
    fecha_registro = Column(DateTime, default=datetime.utcnow)
    observaciones = Column(Text, nullable=True)

    lider = relationship("Lider", back_populates="sufragantes")


class CargaMasivaIncidencia(Base):
    __tablename__ = "carga_masiva_incidencias"

    id = Column(Integer, primary_key=True, index=True)
    lider_id = Column(Integer, ForeignKey("lideres.id"), index=True, nullable=False)
    usuario = Column(String(100), nullable=True)
    archivo = Column(String(255), nullable=True)
    created = Column(Integer, default=0)
    total_rows = Column(Integer, default=0)
    errores_json = Column(Text, nullable=False)  # JSON array de strings: ["Fila 2: ...", ...]
    fecha = Column(DateTime, default=datetime.utcnow, index=True)

    lider = relationship("Lider", back_populates="incidencias_carga_masiva")

Base.metadata.create_all(bind=engine)

# Migración: columna observaciones (si la tabla ya existía)
try:
    with engine.connect() as conn:
        conn.execute(text("ALTER TABLE sufragantes ADD COLUMN observaciones TEXT"))
        conn.commit()
except Exception:
    pass  # Columna ya existe o BD sin tabla sufragantes

# =====================
# Pydantic Models
# =====================

class Token(BaseModel):
    access_token: str
    token_type: str
    user_role: str
    username: str

class TokenData(BaseModel):
    username: Optional[str] = None
    rol: Optional[str] = None

class UsuarioCreate(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=6)
    rol: str = Field(..., pattern="^(superadmin|operador)$")
    email: Optional[str] = Field(None, max_length=255)

class UsuarioResponse(BaseModel):
    id: int
    username: str
    rol: str
    activo: bool
    fecha_creacion: datetime
    ultimo_login: Optional[datetime] = None
    email: Optional[str] = None

    class Config:
        from_attributes = True

class UsuarioUpdateActivo(BaseModel):
    activo: Optional[bool] = None
    email: Optional[str] = Field(None, max_length=255)

class ForgotPasswordRequest(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)

class ResetPasswordRequest(BaseModel):
    token: str = Field(..., min_length=32)
    new_password: str = Field(..., min_length=6)

class LiderCreate(BaseModel):
    nombre: str = Field(..., min_length=2, max_length=200)
    cedula: str = Field(..., min_length=6, max_length=10, pattern="^[0-9]+$")
    edad: int = Field(..., ge=18, le=120)
    celular: str = Field(..., min_length=10, max_length=10, pattern="^[0-9]+$")
    direccion: str = Field(..., min_length=5, max_length=500)
    genero: str = Field(..., pattern="^(M|F|Otro)$")
    departamento: str = Field(..., min_length=2, max_length=100)
    municipio: str = Field(..., min_length=2, max_length=100)
    barrio: Optional[str] = None
    zona_influencia: Optional[str] = None
    tipo_liderazgo: Optional[str] = Field(None, pattern="^(Comunitario|Social|Politico|Religioso|Juvenil|Otro)$")

class LiderResponse(BaseModel):
    id: int
    nombre: str
    cedula: str
    edad: int
    celular: str
    direccion: str
    genero: str
    departamento: str
    municipio: str
    barrio: Optional[str]
    zona_influencia: Optional[str]
    tipo_liderazgo: Optional[str]
    usuario_registro: Optional[str]
    fecha_registro: datetime
    activo: bool

    class Config:
        from_attributes = True

class VerifikResponse(BaseModel):
    """Respuesta de la API de Verifik"""
    success: bool
    data: Optional[dict] = None
    error: Optional[str] = None

class VerifyRequest(BaseModel):
    """Datos ingresados manualmente para comparar con Verifik"""
    cedula: str = Field(..., min_length=6, max_length=10, pattern="^[0-9]+$")
    department: Optional[str] = None
    municipality: Optional[str] = None
    votingStation: Optional[str] = None
    pollingTable: Optional[str] = None
    address: Optional[str] = None

class SufraganteCreate(BaseModel):
    """Modelo para crear sufragante. Solo nombre y cédula obligatorios; el resto opcional (vacío → null o default)."""
    nombre: str = Field(..., min_length=2, max_length=200)
    cedula: str = Field(..., min_length=6, max_length=10, pattern="^[0-9]+$")
    edad: Optional[int] = Field(None, ge=18, le=120)  # Si no se envía, backend usa 18
    celular: Optional[str] = None  # Opcional: null o vacío = "No tiene"; si se envía debe ser 10 dígitos y empezar por 3
    direccion_residencia: Optional[str] = Field(None, max_length=500)  # Si vacío, backend usa "Por definir"
    genero: Optional[str] = Field(None, pattern="^(M|F|Otro)$")
    lider_id: Optional[int] = None
    departamento: Optional[str] = None
    municipio: Optional[str] = None
    lugar_votacion: Optional[str] = None
    mesa_votacion: Optional[str] = None
    direccion_puesto: Optional[str] = None
    estado_validacion: str = Field(default="sin_verificar", pattern="^(verificado|revision|inconsistente|sin_verificar)$")
    discrepancias: Optional[List[str]] = None

    @field_validator("celular")
    @classmethod
    def celular_debe_iniciar_por_3(cls, v):
        if v is None or (isinstance(v, str) and v.strip() in ("", "NO TIENE", "NO TIENE CELULAR")):
            return None
        v = v.strip()
        if not v:
            return None
        if len(v) != 10 or not v.isdigit():
            raise ValueError("El celular debe tener 10 dígitos")
        if v[0] != "3":
            raise ValueError("El celular debe iniciar por el número 3")
        return v

    @field_validator("genero", mode="before")
    @classmethod
    def genero_vacio_a_none(cls, v):
        if v is None or (isinstance(v, str) and not v.strip()):
            return None
        return v.strip() if isinstance(v, str) else v

class SufraganteResponse(BaseModel):
    id: int
    nombre: str
    cedula: str
    edad: int
    celular: Optional[str] = None
    direccion_residencia: str
    genero: Optional[str] = None
    departamento: Optional[str]
    municipio: Optional[str]
    lugar_votacion: Optional[str]
    mesa_votacion: Optional[str]
    direccion_puesto: Optional[str]
    estado_validacion: str
    discrepancias_verifik: Optional[str] = None
    lider_id: Optional[int]
    usuario_registro: Optional[str]
    fecha_registro: datetime
    observaciones: Optional[str] = None

    class Config:
        from_attributes = True

class SufraganteUpdate(BaseModel):
    """Actualización parcial de sufragante (datos Verifik, estado, datos personales; nombre/cedula solo para sin_verificar e inconsistente)"""
    nombre: Optional[str] = Field(None, min_length=2, max_length=200)
    cedula: Optional[str] = Field(None, min_length=6, max_length=10, pattern="^[0-9]+$")
    genero: Optional[str] = Field(None, pattern="^(M|F|Otro)$")
    celular: Optional[str] = None  # 10 dígitos empezando por 3, o vacío/null = "No tiene"
    direccion_residencia: Optional[str] = Field(None, min_length=5, max_length=500)
    departamento: Optional[str] = None
    municipio: Optional[str] = None
    lugar_votacion: Optional[str] = None
    mesa_votacion: Optional[str] = None
    direccion_puesto: Optional[str] = None
    estado_validacion: Optional[str] = Field(None, pattern="^(verificado|revision|inconsistente|sin_verificar)$")
    discrepancias: Optional[List[str]] = None
    observaciones: Optional[str] = None

class DashboardMetrics(BaseModel):
    total_lideres: int
    total_sufragantes: int
    sufragantes_verificados: int
    sufragantes_inconsistentes: int
    sufragantes_en_revision: int
    sufragantes_sin_verificar: int
    registros_hoy: int
    registros_semana: int
    registros_mes: int

class SufraganteMunicipio(BaseModel):
    municipio: str
    total: int

class VotersListResponse(BaseModel):
    """Respuesta paginada de lista de sufragantes."""
    items: List[SufraganteResponse]
    total: int

class SufragantePorLider(BaseModel):
    lider_id: int
    lider_nombre: str
    lider_cedula: str
    total_sufragantes: int
    verificados: int = 0
    sin_verificar: int = 0
    en_revision: int = 0
    inconsistentes: int = 0

class ExportRequest(BaseModel):
    formato: str = Field(..., pattern="^(xlsx)$")

# =====================
# AUTENTICACIÓN (bcrypt directo; passlib incompatible con bcrypt 4.1+)
# =====================

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")

def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return bcrypt.checkpw(
            plain_password.encode("utf-8"),
            hashed_password.encode("utf-8") if isinstance(hashed_password, str) else hashed_password,
        )
    except Exception:
        return False

def get_password_hash(password: str) -> str:
    return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=12)).decode("utf-8")

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def get_current_user(token: str = Depends(oauth2_scheme)) -> TokenData:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="No se pudo validar las credenciales",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        rol: str = payload.get("rol")
        if username is None:
            raise credentials_exception
        return TokenData(username=username, rol=rol)
    except JWTError:
        raise credentials_exception

def require_superadmin(current_user: TokenData = Depends(get_current_user)):
    if current_user.rol != "superadmin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Se requiere rol de superadmin"
        )
    return current_user

# =====================
# DEPENDENCIAS
# =====================

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def normalizar_texto(texto: str) -> str:
    """Normaliza texto a mayúsculas y elimina espacios dobles"""
    if texto:
        return " ".join(texto.upper().split())
    return texto

def normalizar_genero(g: Optional[str]) -> Optional[str]:
    """Normaliza género para BD; None si no se envía (registro masivo)."""
    if g is None or (isinstance(g, str) and not g.strip()):
        return None
    g = g.strip()
    if g.lower() == "otro":
        return "Otro"
    return g.upper() if g in ("M", "F") else None

# =====================
# API VERIFIK
# =====================

async def verificar_cedula(cedula: str) -> dict:
    """
    Consulta la API de Verifik para verificar una cédula
    """
    token = (VERIFIK_TOKEN or "").strip()
    if not token or token == "tu_token_verifik" or len(token) < 20:
        logger.warning("VERIFIK_TOKEN no configurado o inválido (longitud < 20)")
        return {
            "success": False,
            "data": None,
            "error": "API Verifik no configurada. Configure VERIFIK_TOKEN en .env con un token válido de https://verifik.co"
        }

    headers = {
        "Accept": "application/json",
        "Authorization": f"Bearer {token}"
    }

    # Verifik espera el parámetro "documentNumber" (no "cedula"); sin espacios ni puntos
    document_number = (cedula or "").strip().replace(" ", "").replace(".", "")
    params = {"documentNumber": document_number}

    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            response = await client.get(VERIFIK_API_URL, headers=headers, params=params)

        if response.status_code == 200:
            raw = response.json()
            # Verifik devuelve { "data": { department, municipality, address, ... }, "signature": ... }
            data = raw.get("data", raw) if isinstance(raw, dict) else raw
            return {
                "success": True,
                "data": data,
                "error": None
            }
        elif response.status_code == 404:
            return {
                "success": False,
                "data": None,
                "error": "Cédula no encontrada"
            }
        elif response.status_code == 401:
            logger.error(f"Verifik 401 - Token inválido o expirado. Response: {response.text}")
            return {
                "success": False,
                "data": None,
                "error": "Token Verifik inválido o expirado. Renueve VERIFIK_TOKEN en .env desde el panel de Verifik."
            }
        elif response.status_code == 409:
            logger.warning(f"Verifik 409 - Parámetro documentNumber faltante o inválido. Response: {response.text}")
            return {
                "success": False,
                "data": None,
                "error": "Número de cédula requerido. Ingrese una cédula válida (solo números, 6-10 dígitos)."
            }
        else:
            error_msg = f"Error en API Verifik: {response.status_code}"
            logger.error(f"{error_msg} - Response: {response.text}")
            return {
                "success": False,
                "data": None,
                "error": error_msg
            }

    except httpx.RequestError as e:
        logger.error(f"Error de conexión con Verifik: {e}")
        return {
            "success": False,
            "data": None,
            "error": f"Error de conexión: {str(e)}"
        }

def _normalizar_para_comparar(val: Optional[str]) -> str:
    """Normaliza un valor para comparación (strip, mayúsculas, espacios colapsados)."""
    if val is None:
        return ""
    return " ".join(str(val).strip().upper().split())


def _normalizar_lugar_votacion(val: Optional[str]) -> str:
    """
    Normalización flexible solo para 'Lugar de votación':
    quita puntuación, expande abreviaturas, quita sufijos y palabras vacías.
    """
    if val is None:
        return ""
    s = str(val).strip().upper()
    for c in ".-,_;:":
        s = s.replace(c, " ")
    s = " ".join(s.split())
    # Abreviaturas comunes en registraduría
    reemplazos = [
        ("LIC ", "LICEO "), ("LIC.", "LICEO "), ("COL ", "COLEGIO "), ("COL.", "COLEGIO "),
        ("INST ", "INSTITUTO "), ("INST.", "INSTITUTO "), ("ESC ", "ESCUELA "), ("ESC.", "ESCUELA "),
        ("IED ", "INSTITUCION "), ("I E ", "INSTITUCION "),
    ]
    for a, b in reemplazos:
        s = s.replace(a, b)
    # Quitar sufijos típicos (tras guión o espacio)
    for sufijo in ["BACHILLERATO", "PRIMARIA", "SECUNDARIA", "SEDE", "PRINCIPAL", "CENTRO"]:
        if sufijo in s:
            s = s.replace("-" + sufijo, "").replace(" " + sufijo, "")
    s = " ".join(s.split())
    return s


def _coincide_lugar_votacion(verifik_val: Optional[str], manual_val: Optional[str]) -> bool:
    """
    Comparación flexible para Lugar de votación: normalización + contención o palabras clave.
    """
    v = _normalizar_lugar_votacion(verifik_val)
    m = _normalizar_lugar_votacion(manual_val)
    if not v or not m:
        return v == m
    if v == m:
        return True
    # Contención: si uno está contenido en el otro (ej. "LICEO LA MERCED" en "LICEO DE LA MERCED")
    if m in v or v in m:
        return True
    # Palabras clave: quitar stopwords y ver si las palabras del manual están en Verifik
    stop = {"DE", "LA", "EL", "DEL", "LOS", "LAS", "UN", "UNA", "Y", "E", "AL", "EN"}
    words_v = set(w for w in v.split() if w and w not in stop and len(w) > 1)
    words_m = set(w for w in m.split() if w and w not in stop and len(w) > 1)
    if not words_m:
        return True
    return words_m <= words_v or words_v <= words_m


def comparar_datos_verifik(verifik_data: Optional[dict], manual: dict) -> tuple[str, list[str]]:
    """
    Compara datos ingresados manualmente con la respuesta de Verifik.
    Retorna (estado, lista_de_campos_que_no_coinciden).
    Para Lugar de votación se usa comparación flexible (abreviaturas, contención).
    """
    if not verifik_data or not isinstance(verifik_data, dict):
        return "inconsistente", []

    campos = [
        ("department", "departamento"),
        ("municipality", "municipio"),
        ("votingStation", "lugar_votacion"),
        ("pollingTable", "mesa_votacion"),
    ]
    discrepancias = []
    for key_verifik, key_manual in campos:
        v_raw = verifik_data.get(key_verifik)
        m_raw = manual.get(key_verifik) or manual.get(key_manual)
        if key_verifik == "votingStation":
            coincide = _coincide_lugar_votacion(v_raw, m_raw)
        else:
            coincide = _normalizar_para_comparar(v_raw) == _normalizar_para_comparar(m_raw)
        if not coincide:
            discrepancias.append(key_verifik)

    if not discrepancias:
        return "verificado", []
    return "revision", discrepancias

# =====================
# APLICACIÓN FASTAPI
# =====================

app = FastAPI(
    title="API Sistema de Registro",
    description="API REST para gestión de líderes y sufragantes",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def startup_add_discrepancias_column():
    """Añade columna discrepancias_verifik si la tabla ya existía sin ella."""
    try:
        with engine.connect() as conn:
            with conn.begin():
                conn.execute(text("ALTER TABLE sufragantes ADD COLUMN IF NOT EXISTS discrepancias_verifik TEXT"))
                try:
                    conn.execute(text("ALTER TABLE sufragantes ALTER COLUMN celular DROP NOT NULL"))
                except Exception:
                    pass
                try:
                    conn.execute(text("ALTER TABLE sufragantes DROP CONSTRAINT IF EXISTS sufragantes_estado_validacion_check"))
                    conn.execute(text("ALTER TABLE sufragantes ADD CONSTRAINT sufragantes_estado_validacion_check CHECK (estado_validacion IN ('verificado', 'inconsistente', 'revision', 'sin_verificar'))"))
                except Exception:
                    pass
                conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS reset_token VARCHAR(64)"))
                conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS reset_token_expires TIMESTAMP"))
                conn.execute(text("ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS email VARCHAR(255)"))
            try:
                conn.execute(text("CREATE INDEX IF NOT EXISTS ix_usuarios_reset_token ON usuarios(reset_token)"))
            except Exception:
                pass
            try:
                with conn.begin():
                    # Buscar el nombre real del constraint CHECK de genero (puede variar según cómo se creó la tabla)
                    r = conn.execute(text("""
                        SELECT c.conname FROM pg_constraint c
                        JOIN pg_attribute a ON a.attrelid = c.conrelid AND a.attnum = ANY(c.conkey) AND NOT a.attisdropped
                        WHERE c.conrelid = 'sufragantes'::regclass AND c.contype = 'c' AND a.attname = 'genero'
                    """))
                    names = [row[0] for row in r]
                    for name in names:
                        conn.execute(text(f'ALTER TABLE sufragantes DROP CONSTRAINT IF EXISTS "{name}"'))
                    conn.execute(text("ALTER TABLE sufragantes DROP CONSTRAINT IF EXISTS sufragantes_genero_check"))
                    conn.execute(text("ALTER TABLE sufragantes ALTER COLUMN genero DROP NOT NULL"))
                    conn.execute(text("ALTER TABLE sufragantes ADD CONSTRAINT sufragantes_genero_check CHECK (genero IS NULL OR genero IN ('M', 'F', 'Otro'))"))
                logger.info("Migración: columna sufragantes.genero permitida como NULL")
            except Exception as e:
                logger.warning(f"Migración genero (puede estar ya aplicada): {e}")
    except Exception as e:
        logger.warning(f"Startup: no se pudo añadir columnas (puede existir ya): {e}")

# =====================
# RUTAS DE AUTENTICACIÓN
# =====================

@app.post("/auth/login", response_model=Token)
async def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """
    Endpoint de login para usuarios
    """
    usuario = db.query(Usuario).filter(Usuario.username == form_data.username).first()

    if not usuario:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario no encontrado",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if not usuario.activo:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Usuario inactivo",
            headers={"WWW-Authenticate": "Bearer"},
        )

    if not verify_password(form_data.password, usuario.password_hash):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Contraseña incorrecta",
            headers={"WWW-Authenticate": "Bearer"},
        )

    # Actualizar último login
    usuario.ultimo_login = datetime.utcnow()
    db.commit()

    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": usuario.username, "rol": usuario.rol},
        expires_delta=access_token_expires
    )

    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user_role": usuario.rol,
        "username": usuario.username
    }

@app.post("/auth/register", response_model=UsuarioResponse)
async def register(usuario_data: UsuarioCreate,
                  current_user: TokenData = Depends(require_superadmin),
                  db: Session = Depends(get_db)):
    """
    Registrar nuevo usuario (solo superadmin)
    """
    # Verificar si ya existe el username
    existente = db.query(Usuario).filter(Usuario.username == usuario_data.username).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El nombre de usuario ya existe"
        )

    # Crear nuevo usuario
    nuevo_usuario = Usuario(
        username=usuario_data.username,
        password_hash=get_password_hash(usuario_data.password),
        rol=usuario_data.rol,
        activo=True,
        email=usuario_data.email.strip() if usuario_data.email and usuario_data.email.strip() else None
    )

    db.add(nuevo_usuario)
    db.commit()
    db.refresh(nuevo_usuario)

    logger.info(f"Usuario {current_user.username} creó nuevo usuario: {usuario_data.username}")

    return nuevo_usuario

@app.get("/auth/me", response_model=UsuarioResponse)
async def get_current_user_info(current_user: TokenData = Depends(get_current_user),
                                db: Session = Depends(get_db)):
    """
    Obtener información del usuario actual
    """
    usuario = db.query(Usuario).filter(Usuario.username == current_user.username).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return usuario

RESET_TOKEN_EXPIRE_HOURS = 1

def _send_reset_password_email(to_email: str, reset_link: str, username: str) -> bool:
    """Envía correo con el enlace de restablecimiento de contraseña (Gmail SMTP)."""
    if not GMAIL_USER or not GMAIL_APP_PASSWORD:
        logger.warning("GMAIL_USER o GMAIL_APP_PASSWORD no configurados; no se envía correo.")
        return False
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = "InnovaBigData - Restablecer contraseña"
        msg["From"] = f"InnovaBigData <{GMAIL_USER}>"
        msg["To"] = to_email
        text = f"""Hola,\n\nHas solicitado restablecer la contraseña del usuario {username} en InnovaBigData.\n\nHaz clic en el siguiente enlace (válido 1 hora):\n{reset_link}\n\nSi no solicitaste este correo, ignóralo.\n\n— InnovaBigData"""
        html = f"""<p>Hola,</p><p>Has solicitado restablecer la contraseña del usuario <strong>{username}</strong> en InnovaBigData.</p><p><a href="{reset_link}">Haz clic aquí para restablecer tu contraseña</a> (válido 1 hora).</p><p>Si no solicitaste este correo, ignóralo.</p><p>— InnovaBigData</p>"""
        msg.attach(MIMEText(text, "plain", "utf-8"))
        msg.attach(MIMEText(html, "html", "utf-8"))
        with smtplib.SMTP("smtp.gmail.com", 587) as server:
            server.starttls()
            server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
            server.sendmail(GMAIL_USER, to_email, msg.as_string())
        logger.info(f"Correo de recuperación enviado a {to_email}")
        return True
    except Exception as e:
        logger.exception(f"Error enviando correo de recuperación a {to_email}: {e}")
        return False

@app.post("/auth/forgot-password")
async def forgot_password(body: ForgotPasswordRequest, db: Session = Depends(get_db)):
    """
    Solicitar restablecimiento de contraseña por correo. Disponible para operadores y superadmin:
    se envía un enlace al email registrado del usuario (campo email en la tabla usuarios).
    """
    usuario = db.query(Usuario).filter(Usuario.username == body.username.strip()).first()
    if not usuario:
        return {"success": True, "message": "Si el usuario existe y tiene correo registrado, recibirá un enlace en unos minutos. Revise su bandeja de entrada y spam."}
    if not usuario.activo:
        return {"success": True, "message": "Si el usuario existe y tiene correo registrado, recibirá un enlace en unos minutos. Revise su bandeja de entrada y spam."}
    email = (usuario.email or "").strip()
    if not email:
        return {"success": True, "message": "Este usuario no tiene correo registrado. Contacte al administrador para que asigne un correo y pueda recuperar la contraseña."}
    token = secrets.token_urlsafe(32)
    usuario.reset_token = token
    usuario.reset_token_expires = datetime.utcnow() + timedelta(hours=RESET_TOKEN_EXPIRE_HOURS)
    db.commit()
    reset_link = f"{RESET_PASSWORD_BASE_URL}/reset-password?token={quote_plus(token)}"
    sent = _send_reset_password_email(email, reset_link, usuario.username)
    if not sent:
        return {"success": False, "message": "No se pudo enviar el correo. Intente más tarde o contacte al administrador."}
    return {
        "success": True,
        "message": "Si el usuario existe y tiene correo registrado, recibirá un enlace en unos minutos. Revise su bandeja de entrada y spam."
    }

@app.post("/auth/reset-password")
async def reset_password(body: ResetPasswordRequest, db: Session = Depends(get_db)):
    """
    Restablecer contraseña con el token recibido en forgot-password.
    """
    usuario = db.query(Usuario).filter(
        Usuario.reset_token == body.token,
        Usuario.reset_token_expires > datetime.utcnow()
    ).first()
    if not usuario:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Enlace inválido o expirado. Solicite uno nuevo desde «Olvidé mi contraseña»."
        )
    usuario.password_hash = get_password_hash(body.new_password)
    usuario.reset_token = None
    usuario.reset_token_expires = None
    db.commit()
    return {"success": True, "message": "Contraseña actualizada. Ya puede iniciar sesión."}

@app.get("/users", response_model=List[UsuarioResponse])
async def listar_usuarios(
    rol: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Listar usuarios (solo superadmin). Por defecto lista operadores; rol=superadmin para ver todos.
    """
    query = db.query(Usuario)
    if rol:
        query = query.filter(Usuario.rol == rol)
    else:
        query = query.filter(Usuario.rol == "operador")
    usuarios = query.order_by(Usuario.fecha_creacion.desc()).all()
    return usuarios

@app.patch("/users/{user_id}", response_model=UsuarioResponse)
async def actualizar_usuario_activo(
    user_id: int,
    data: UsuarioUpdateActivo,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Activar o desactivar un usuario (solo superadmin). No se puede desactivar a sí mismo.
    """
    usuario = db.query(Usuario).filter(Usuario.id == user_id).first()
    if not usuario:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if usuario.username == current_user.username:
        raise HTTPException(status_code=400, detail="No puede desactivar su propio usuario")
    if data.activo is not None:
        usuario.activo = data.activo
    if data.email is not None:
        usuario.email = data.email.strip() or None
    db.commit()
    db.refresh(usuario)
    logger.info(f"Usuario {current_user.username} actualizó usuario {usuario.username}")
    return usuario

# =====================
# RUTAS DE LÍDERES
# =====================

@app.get("/leaders", response_model=List[LiderResponse])
async def listar_lideres(
    skip: int = 0,
    limit: int = 100,
    activo: Optional[bool] = None,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Listar líderes con paginación y filtros
    """
    query = db.query(Lider)

    if activo is not None:
        query = query.filter(Lider.activo == activo)

    lideres = query.order_by(Lider.nombre).offset(skip).limit(limit).all()
    return lideres

@app.get("/leaders/all", response_model=List[LiderResponse])
async def listar_todos_lideres(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Listar todos los líderes activos (para selects)
    """
    lideres = db.query(Lider).filter(Lider.activo == True).order_by(Lider.nombre).all()
    return lideres

@app.post("/leaders", response_model=LiderResponse)
async def crear_lider(
    lider_data: LiderCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Crear nuevo líder (solo superadmin, sin verificación Verifik)
    """
    # Verificar cédula única
    existente = db.query(Lider).filter(Lider.cedula == lider_data.cedula).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cédula ya está registrada"
        )

    # Normalizar datos
    nuevo_lider = Lider(
        nombre=normalizar_texto(lider_data.nombre),
        cedula=lider_data.cedula,
        edad=lider_data.edad,
        celular=lider_data.celular,
        direccion=lider_data.direccion,
        genero=lider_data.genero.upper(),
        departamento=lider_data.departamento.upper(),
        municipio=lider_data.municipio.upper(),
        barrio=normalizar_texto(lider_data.barrio),
        zona_influencia=lider_data.zona_influencia,
        tipo_liderazgo=lider_data.tipo_liderazgo,
        usuario_registro=current_user.username
    )

    db.add(nuevo_lider)
    db.commit()
    db.refresh(nuevo_lider)

    logger.info(f"Superadmin {current_user.username} creó líder: {nuevo_lider.id}")

    return nuevo_lider

@app.put("/leaders/{lider_id}", response_model=LiderResponse)
async def actualizar_lider(
    lider_id: int,
    lider_data: LiderCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Actualizar líder (solo superadmin)
    """
    lider = db.query(Lider).filter(Lider.id == lider_id).first()
    if not lider:
        raise HTTPException(status_code=404, detail="Líder no encontrado")

    # Verificar cédula única (si cambió)
    if lider_data.cedula != lider.cedula:
        existente = db.query(Lider).filter(Lider.cedula == lider_data.cedula).first()
        if existente:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La cédula ya está registrada"
            )

    # Actualizar campos
    lider.nombre = normalizar_texto(lider_data.nombre)
    lider.cedula = lider_data.cedula
    lider.edad = lider_data.edad
    lider.celular = lider_data.celular
    lider.direccion = lider_data.direccion
    lider.genero = lider_data.genero.upper()
    lider.departamento = lider_data.departamento.upper()
    lider.municipio = lider_data.municipio.upper()
    lider.barrio = normalizar_texto(lider_data.barrio)
    lider.zona_influencia = lider_data.zona_influencia
    lider.tipo_liderazgo = lider_data.tipo_liderazgo

    db.commit()
    db.refresh(lider)

    logger.info(f"Superadmin {current_user.username} actualizó líder: {lider_id}")

    return lider

@app.delete("/leaders/{lider_id}")
async def desactivar_lider(
    lider_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Desactivar líder (solo superadmin)
    """
    lider = db.query(Lider).filter(Lider.id == lider_id).first()
    if not lider:
        raise HTTPException(status_code=404, detail="Líder no encontrado")

    lider.activo = False
    db.commit()

    logger.info(f"Superadmin {current_user.username} desactivó líder: {lider_id}")

    return {"message": "Líder desactivado correctamente"}

# =====================
# RUTAS DE SUFRAGANTES
# =====================

@app.post("/voters/verify")
async def verificar_cedula_endpoint(
    body: Optional[VerifyRequest] = None,
    cedula: Optional[str] = Query(None, alias="cedula"),
    exclude_voter_id: Optional[int] = Query(None, description="Al re-verificar un sufragante existente, pasar su ID para no considerarlo duplicado"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Verifica cédula en Verifik y compara con datos ingresados manualmente.
    Si se envían datos manuales (body), retorna estado: verificado | revision | inconsistente.
    Para re-verificar un sufragante ya registrado, enviar exclude_voter_id con el id del sufragante.
    """
    doc = (body.cedula if body else cedula) or ""
    if not doc or not doc.isdigit() or len(doc) < 6 or len(doc) > 10:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cédula inválida: debe tener entre 6 y 10 dígitos"
        )

    q = db.query(Sufragante).filter(Sufragante.cedula == doc)
    if exclude_voter_id is not None:
        q = q.filter(Sufragante.id != exclude_voter_id)
    existente = q.first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cédula ya está registrada en el sistema"
        )

    resultado = await verificar_cedula(doc)
    verifik_data = resultado.get("data") if resultado["success"] else None

    if not resultado["success"]:
        return {
            "success": False,
            "error": resultado["error"],
            "data": None,
            "estado": "inconsistente",
            "discrepancias": []
        }

    if not body:
        return {
            "success": True,
            "data": verifik_data,
            "error": None,
            "estado": None,
            "discrepancias": []
        }

    manual = {
        "department": body.department or "",
        "municipality": body.municipality or "",
        "votingStation": body.votingStation or "",
        "pollingTable": body.pollingTable or "",
        "address": body.address or "",
    }
    estado, discrepancias = comparar_datos_verifik(verifik_data, manual)

    return {
        "success": True,
        "data": verifik_data,
        "error": None,
        "estado": estado,
        "discrepancias": discrepancias
    }

@app.post("/voters", response_model=SufraganteResponse)
async def crear_sufragante(
    sufragante_data: SufraganteCreate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Registrar nuevo sufragante. Los datos de Verifik y estado se envían desde el formulario
    (tras haber hecho "Verificar"). Permite registrar aunque el estado sea inconsistente.
    """
    nombre_normalizado = normalizar_texto(sufragante_data.nombre)

    existente = db.query(Sufragante).filter(Sufragante.cedula == sufragante_data.cedula).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La cédula ya está registrada"
        )

    if sufragante_data.lider_id:
        lider = db.query(Lider).filter(Lider.id == sufragante_data.lider_id, Lider.activo == True).first()
        if not lider:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El líder especificado no existe o está inactivo"
            )

    discrepancias_json = None
    if sufragante_data.discrepancias and sufragante_data.estado_validacion == "revision":
        discrepancias_json = json.dumps(sufragante_data.discrepancias)

    celular_val = None
    if sufragante_data.celular and str(sufragante_data.celular).strip().upper() not in ("", "NO TIENE", "NO TIENE CELULAR"):
        celular_val = sufragante_data.celular.strip()

    edad_val = sufragante_data.edad if sufragante_data.edad is not None else 18
    dir_val = (sufragante_data.direccion_residencia or "").strip()
    if not dir_val:
        dir_val = "Por definir"

    nuevo_sufragante = Sufragante(
        nombre=nombre_normalizado,
        cedula=sufragante_data.cedula,
        edad=edad_val,
        celular=celular_val,
        direccion_residencia=dir_val[:500],
        genero=normalizar_genero(sufragante_data.genero),
        estado_validacion=sufragante_data.estado_validacion,
        discrepancias_verifik=discrepancias_json,
        lider_id=sufragante_data.lider_id,
        usuario_registro=current_user.username,
        departamento=normalizar_texto(sufragante_data.departamento) if sufragante_data.departamento else None,
        municipio=normalizar_texto(sufragante_data.municipio) if sufragante_data.municipio else None,
        lugar_votacion=normalizar_texto(sufragante_data.lugar_votacion) if sufragante_data.lugar_votacion else None,
        mesa_votacion=normalizar_texto(sufragante_data.mesa_votacion) if sufragante_data.mesa_votacion else None,
        direccion_puesto=sufragante_data.direccion_puesto or None
    )

    db.add(nuevo_sufragante)
    db.commit()
    db.refresh(nuevo_sufragante)

    logger.info(f"Usuario {current_user.username} registró sufragante: {nuevo_sufragante.id}")

    return nuevo_sufragante

@app.get("/voters", response_model=VotersListResponse)
async def listar_sufragantes(
    skip: int = 0,
    limit: int = 20,
    lider_id: Optional[int] = None,
    estado: Optional[str] = None,
    municipio: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Listar sufragantes con filtros y paginación (items + total).
    """
    query = db.query(Sufragante)

    if lider_id:
        query = query.filter(Sufragante.lider_id == lider_id)
    if estado:
        query = query.filter(Sufragante.estado_validacion == estado)
    if municipio:
        query = query.filter(Sufragante.municipio == municipio.upper())

    # Los operadores solo ven registros que ellos crearon
    if current_user.rol == "operador":
        query = query.filter(Sufragante.usuario_registro == current_user.username)

    total = query.count()
    sufragantes = query.order_by(Sufragante.fecha_registro.desc()).offset(skip).limit(limit).all()
    return VotersListResponse(items=sufragantes, total=total)

@app.get("/voters/{voter_id}", response_model=SufraganteResponse)
async def obtener_sufragante(
    voter_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Obtener detalles de un sufragante
    """
    sufragante = db.query(Sufragante).filter(Sufragante.id == voter_id).first()
    if not sufragante:
        raise HTTPException(status_code=404, detail="Sufragante no encontrado")

    # Los operadores solo pueden ver sus propios registros
    if current_user.rol == "operador" and sufragante.usuario_registro != current_user.username:
        raise HTTPException(status_code=403, detail="No tiene acceso a este registro")

    return sufragante

@app.patch("/voters/{voter_id}", response_model=SufraganteResponse)
async def actualizar_sufragante(
    voter_id: int,
    data: SufraganteUpdate,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Actualizar datos de un sufragante (datos Verifik y/o estado de verificación).
    Útil para corregir información en sufragantes "En revisión" y actualizar estado tras verificar.
    """
    sufragante = db.query(Sufragante).filter(Sufragante.id == voter_id).first()
    if not sufragante:
        raise HTTPException(status_code=404, detail="Sufragante no encontrado")

    if current_user.rol == "operador" and sufragante.usuario_registro != current_user.username:
        raise HTTPException(status_code=403, detail="No tiene acceso a este registro")

    update = data.model_dump(exclude_unset=True)
    if "nombre" in update and update["nombre"]:
        sufragante.nombre = normalizar_texto(update["nombre"])
    if "cedula" in update and update["cedula"]:
        nueva_cedula = update["cedula"].strip()
        if len(nueva_cedula) < 6 or len(nueva_cedula) > 10 or not nueva_cedula.isdigit():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Cédula inválida: entre 6 y 10 dígitos")
        otro = db.query(Sufragante).filter(Sufragante.cedula == nueva_cedula, Sufragante.id != voter_id).first()
        if otro:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="La cédula ya está registrada por otro sufragante")
        sufragante.cedula = nueva_cedula
    if "departamento" in update:
        sufragante.departamento = normalizar_texto(update["departamento"]) if update["departamento"] else None
    if "municipio" in update:
        sufragante.municipio = normalizar_texto(update["municipio"]) if update["municipio"] else None
    if "lugar_votacion" in update:
        sufragante.lugar_votacion = normalizar_texto(update["lugar_votacion"]) if update["lugar_votacion"] else None
    if "mesa_votacion" in update:
        sufragante.mesa_votacion = normalizar_texto(update["mesa_votacion"]) if update["mesa_votacion"] else None
    if "direccion_puesto" in update:
        sufragante.direccion_puesto = update["direccion_puesto"] or None
    if "estado_validacion" in update:
        sufragante.estado_validacion = update["estado_validacion"]
    if "discrepancias" in update:
        sufragante.discrepancias_verifik = json.dumps(update["discrepancias"]) if update["discrepancias"] else None
    if "genero" in update:
        sufragante.genero = normalizar_genero(update.get("genero"))
    if "celular" in update:
        cel = (update["celular"] or "").strip()
        if not cel or cel.upper() in ("NO TIENE", "NO TIENE CELULAR"):
            sufragante.celular = None
        else:
            cel = cel.replace(" ", "")
            if len(cel) != 10 or not cel.isdigit() or cel[0] != "3":
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Celular debe tener 10 dígitos e iniciar por 3 (o vacío)")
            sufragante.celular = cel
    if "direccion_residencia" in update and update["direccion_residencia"]:
        sufragante.direccion_residencia = normalizar_texto(update["direccion_residencia"])[:500]
    if "observaciones" in update:
        sufragante.observaciones = (update["observaciones"] or "").strip() or None

    db.commit()
    db.refresh(sufragante)
    logger.info(f"Usuario {current_user.username} actualizó sufragante {sufragante.id}")
    return sufragante


@app.delete("/voters/{voter_id}")
async def eliminar_sufragante(
    voter_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Eliminar un sufragante por ID (solo superadmin).
    """
    sufragante = db.query(Sufragante).filter(Sufragante.id == voter_id).first()
    if not sufragante:
        raise HTTPException(status_code=404, detail="Sufragante no encontrado")

    db.delete(sufragante)
    db.commit()
    logger.info(f"Superadmin {current_user.username} eliminó sufragante {voter_id}")
    return {"message": "Sufragante eliminado", "deleted_id": voter_id}


@app.delete("/voters/by-leader/{lider_id}")
async def eliminar_sufragantes_por_lider(
    lider_id: int,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Eliminar todos los sufragantes asociados a un líder (solo superadmin).
    """
    deleted = db.query(Sufragante).filter(Sufragante.lider_id == lider_id).delete(synchronize_session=False)
    db.commit()
    logger.info(f"Superadmin {current_user.username} eliminó {deleted} sufragante(s) del líder {lider_id}")
    return {"message": "Sufragantes eliminados", "lider_id": lider_id, "deleted": deleted}

def _normalize_header(h: str) -> str:
    """Normaliza nombre de columna para comparación (sin acentos, mayúsculas, sin espacios dobles)."""
    if h is None:
        return ""
    s = str(h).strip().upper().replace("  ", " ").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Á", "A")
    return s

def _col_index(header_row: list, names: list, default: int) -> int:
    """Devuelve el índice de la primera columna cuyo encabezado coincida con alguno de names."""
    for i, cell in enumerate(header_row):
        if i >= 50:
            break
        h = _normalize_header(cell)
        for n in names:
            if n in h or h in n:
                return i
    return default

def _parse_excel_masivo(contents: bytes) -> List[dict]:
    """
    Parsea Excel de registro masivo. Detecta columnas por nombre de encabezado (primera fila).
    NOMBRES Y APELLIDOS, CÉDULA, EDAD, CELULAR, DIRECCION, QUIEN REFIERE (se ignora),
    DEPARTAMENTO, MUNICIPIO, LUGAR DE VOTACION, MESA DE VOTACION.
    """
    wb = openpyxl.load_workbook(BytesIO(contents), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_row = list(rows[0]) if rows else []
    # Detección por nombre para soportar columnas en distinto orden o columna extra al inicio
    idx_nombre = _col_index(header_row, ["NOMBRES Y APELLIDOS", "NOMBRE"], 0)
    idx_cedula = _col_index(header_row, ["CEDULA", "CÉDULA"], 1)
    idx_edad = _col_index(header_row, ["EDAD"], 2)
    idx_celular = _col_index(header_row, ["CELULAR"], 3)
    idx_direccion = _col_index(header_row, ["DIRECCION", "DIRECCIÓN"], 4)
    idx_dep = _col_index(header_row, ["DEPARTAMENTO"], 6)
    idx_mun = _col_index(header_row, ["MUNICIPIO"], 7)
    idx_lugar = _col_index(header_row, ["LUGAR DE VOTACION", "LUGAR VOTACION"], 8)
    idx_mesa = _col_index(header_row, ["MESA DE VOTACION", "MESA VOTACION"], 9)
    idx_observaciones = _col_index(header_row, ["OBSERVACIONES"], -1)
    out = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if not row or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in row):
            continue
        row = list(row)
        nombre = (row[idx_nombre] or "").strip() if idx_nombre < len(row) else ""
        cedula_raw = row[idx_cedula] if idx_cedula < len(row) else None
        if cedula_raw is not None and isinstance(cedula_raw, (int, float)):
            cedula = str(int(cedula_raw))
        else:
            cedula = (str(cedula_raw or "").strip().replace(" ", "").replace(".", "") if cedula_raw is not None else "") or ""
        edad_val = row[idx_edad] if idx_edad < len(row) else None
        if isinstance(edad_val, float) and edad_val == int(edad_val):
            edad_val = int(edad_val)
        celular_raw = row[idx_celular] if idx_celular < len(row) else None
        if celular_raw is not None and isinstance(celular_raw, (int, float)):
            celular = str(int(celular_raw)) if celular_raw else ""
        else:
            celular = (str(celular_raw or "").strip().replace(" ", "") if celular_raw is not None else "") or ""
        direccion = (str(row[idx_direccion] or "").strip() if idx_direccion < len(row) else "") or ""
        dep = (str(row[idx_dep] or "").strip() if idx_dep < len(row) else "") or ""
        mun = (str(row[idx_mun] or "").strip() if idx_mun < len(row) else "") or ""
        lugar = (str(row[idx_lugar] or "").strip() if idx_lugar < len(row) else "") or ""
        mesa = (str(row[idx_mesa] or "").strip() if idx_mesa < len(row) else "") or ""
        observaciones = (str(row[idx_observaciones] or "").strip() if idx_observaciones >= 0 and idx_observaciones < len(row) else "") or ""
        try:
            edad = int(edad_val) if edad_val is not None else None
        except (TypeError, ValueError):
            edad = None
        out.append({
            "row": row_idx,
            "nombre": nombre,
            "cedula": cedula,
            "edad": edad,
            "celular": celular or None,
            "direccion_residencia": direccion,
            "departamento": dep or None,
            "municipio": mun or None,
            "lugar_votacion": lugar or None,
            "mesa_votacion": mesa or None,
            "observaciones": observaciones or None,
        })
    return out

@app.post("/voters/upload")
async def upload_sufragantes_masivo(
    file: UploadFile = File(...),
    lider_id: int = Form(...),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Registro masivo de sufragantes desde Excel. Seleccionar líder y adjuntar archivo.
    Columnas: NOMBRES Y APELLIDOS, CÉDULA, EDAD, CELULAR, DIRECCION (residencia), DEPARTAMENTO, MUNICIPIO, LUGAR DE VOTACION, MESA DE VOTACION.
    LUGAR DE VOTACION y MESA DE VOTACION pueden ir vacíos. QUIEN REFIERE se ignora. Género queda null. Estado: sin_verificar.
    """
    try:
        if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
            raise HTTPException(status_code=400, detail="Debe adjuntar un archivo Excel (.xlsx o .xls)")
        contents = await file.read()
        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="El archivo no debe superar 10 MB")
        lider = db.query(Lider).filter(Lider.id == lider_id, Lider.activo == True).first()
        if not lider:
            raise HTTPException(status_code=400, detail="El líder seleccionado no existe o está inactivo")
        rows = _parse_excel_masivo(contents)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error en upload masivo (lectura/parseo)")
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")
    if not rows:
        return {"created": 0, "errors": ["El archivo no tiene filas de datos o el formato no es válido."]}
    created = 0
    errors = []
    for r in rows:
        nombre = (r.get("nombre") or "").strip()
        cedula = (r.get("cedula") or "").strip()
        edad_raw = r.get("edad")
        direccion = (r.get("direccion_residencia") or "").strip()
        if not nombre or len(nombre) < 2:
            errors.append(f"Fila {r['row']}: nombre inválido")
            continue
        if not cedula or len(cedula) < 6 or not cedula.isdigit():
            errors.append(f"Fila {r['row']}: cédula inválida (6-10 dígitos)")
            continue
        try:
            edad_num = int(float(edad_raw)) if edad_raw not in (None, "") else None
        except (TypeError, ValueError):
            edad_num = None
        if edad_num is not None and (edad_num < 18 or edad_num > 120):
            errors.append(f"Fila {r['row']}: edad debe estar entre 18 y 120 (o vacío)")
            continue
        edad = edad_num if edad_num is not None and 18 <= edad_num <= 120 else 18
        if not direccion:
            direccion = "Por definir"
        celular = r.get("celular")
        if celular:
            celular = str(celular).strip()
            if len(celular) != 10 or not celular.isdigit() or celular[0] != "3":
                errors.append(f"Fila {r['row']}: celular debe ser 10 dígitos e iniciar por 3 (o vacío)")
                continue
        else:
            celular = None
        if db.query(Sufragante).filter(Sufragante.cedula == cedula).first():
            errors.append(f"Fila {r['row']}: cédula {cedula} ya registrada")
            continue
        nombre_norm = normalizar_texto(nombre)
        # Lugar y mesa pueden venir vacíos en el Excel; guardar como None
        lugar_val = (r.get("lugar_votacion") or "").strip() if r.get("lugar_votacion") is not None else ""
        mesa_val = (r.get("mesa_votacion") or "").strip() if r.get("mesa_votacion") is not None else ""
        try:
            suf = Sufragante(
                nombre=nombre_norm,
                cedula=cedula,
                edad=edad,
                celular=celular or None,
                direccion_residencia=direccion[:500],
                genero=None,
                estado_validacion="sin_verificar",
                discrepancias_verifik=None,
                lider_id=lider_id,
                usuario_registro=current_user.username,
                departamento=normalizar_texto(r["departamento"]) if (r.get("departamento") or "").strip() else None,
                municipio=normalizar_texto(r["municipio"]) if (r.get("municipio") or "").strip() else None,
                lugar_votacion=normalizar_texto(lugar_val) if lugar_val else None,
                mesa_votacion=mesa_val or None,
                direccion_puesto=None,
                observaciones=(r.get("observaciones") or "").strip() or None,
            )
            db.add(suf)
            db.commit()
            created += 1
        except IntegrityError as e:
            db.rollback()
            err_msg = str(e.orig) if getattr(e, "orig", None) else str(e)
            if "cedula" in err_msg.lower() or "unique" in err_msg.lower():
                errors.append(f"Fila {r['row']}: cédula {cedula} ya registrada")
            else:
                errors.append(f"Fila {r['row']}: error de BD - {err_msg[:100]}")
        except Exception as e:
            db.rollback()
            logger.exception("Error creando sufragante en carga masiva")
            errors.append(f"Fila {r['row']}: {str(e)}")
    # Guardar incidencias (errores) asociadas al líder para exportación posterior (solo si hubo errores)
    if errors:
        try:
            inc = CargaMasivaIncidencia(
                lider_id=lider_id,
                usuario=current_user.username,
                archivo=file.filename or None,
                created=created,
                total_rows=len(rows),
                errores_json=json.dumps(errors, ensure_ascii=False),
            )
            db.add(inc)
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("No se pudo guardar incidencias de carga masiva")
    return {"created": created, "errors": errors, "total_rows": len(rows)}

# =====================
# DASHBOARD Y REPORTES
# =====================

@app.get("/dashboard", response_model=DashboardMetrics)
async def get_dashboard(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Obtener métricas del dashboard
    """
    # Contadores principales
    total_lideres = db.query(func.count(Lider.id)).filter(Lider.activo == True).scalar()
    total_sufragantes = db.query(func.count(Sufragante.id)).scalar()

    # Estados de validación
    verificados = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "verificado").scalar()
    inconsistentes = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "inconsistente").scalar()
    en_revision = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "revision").scalar()
    sin_verificar = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "sin_verificar").scalar()

    # Registros por período
    hoy = datetime.utcnow().date()
    semana_inicio = hoy - timedelta(days=hoy.weekday())
    mes_inicio = hoy.replace(day=1)

    registros_hoy = db.query(func.count(Sufragante.id)).filter(
        func.date(Sufragante.fecha_registro) == hoy
    ).scalar()

    registros_semana = db.query(func.count(Sufragante.id)).filter(
        Sufragante.fecha_registro >= semana_inicio
    ).scalar()

    registros_mes = db.query(func.count(Sufragante.id)).filter(
        Sufragante.fecha_registro >= mes_inicio
    ).scalar()

    return DashboardMetrics(
        total_lideres=total_lideres or 0,
        total_sufragantes=total_sufragantes or 0,
        sufragantes_verificados=verificados or 0,
        sufragantes_inconsistentes=inconsistentes or 0,
        sufragantes_en_revision=en_revision or 0,
        sufragantes_sin_verificar=sin_verificar or 0,
        registros_hoy=registros_hoy or 0,
        registros_semana=registros_semana or 0,
        registros_mes=registros_mes or 0
    )

@app.get("/dashboard/municipios", response_model=List[SufraganteMunicipio])
async def get_sufragantes_por_municipio(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Obtener distribución de sufragantes por municipio
    """
    resultados = db.query(
        Sufragante.municipio,
        func.count(Sufragante.id).label("total")
    ).group_by(Sufragante.municipio).having(Sufragante.municipio != None).all()

    return [
        SufraganteMunicipio(municipio=row[0] or "Sin especificar", total=row[1])
        for row in resultados
    ]

@app.get("/dashboard/lideres", response_model=List[SufragantePorLider])
async def get_sufragantes_por_lider(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Obtener conteo de sufragantes por líder con desglose por estado (verificado, sin verificar, en revisión, inconsistente).
    """
    resultados = db.query(
        Lider.id,
        Lider.nombre,
        Lider.cedula,
        func.count(Sufragante.id).label("total"),
        func.count(case((Sufragante.estado_validacion == "verificado", 1), else_=None)).label("verificados"),
        func.count(case((Sufragante.estado_validacion == "sin_verificar", 1), else_=None)).label("sin_verificar"),
        func.count(case((Sufragante.estado_validacion == "revision", 1), else_=None)).label("en_revision"),
        func.count(case((Sufragante.estado_validacion == "inconsistente", 1), else_=None)).label("inconsistentes"),
    ).outerjoin(Sufragante, Lider.id == Sufragante.lider_id).group_by(Lider.id).all()

    return [
        SufragantePorLider(
            lider_id=row[0],
            lider_nombre=row[1],
            lider_cedula=row[2],
            total_sufragantes=row[3] or 0,
            verificados=row[4] or 0,
            sin_verificar=row[5] or 0,
            en_revision=row[6] or 0,
            inconsistentes=row[7] or 0
        )
        for row in resultados
    ]

@app.get("/dashboard/operadores")
async def get_registros_por_operador(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Obtener conteo de registros por operador
    """
    resultados = db.query(
        Sufragante.usuario_registro,
        func.count(Sufragante.id).label("total")
    ).group_by(Sufragante.usuario_registro).all()

    return [{"operador": row[0] or "Sin usuario", "total": row[1]} for row in resultados]

@app.get("/dashboard/tendencia")
async def get_tendencia_registros(
    dias: int = 30,
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Obtener tendencia de registros por día
    """
    from sqlalchemy import text

    resultados = db.query(
        func.date(Sufragante.fecha_registro).label("fecha"),
        func.count(Sufragante.id).label("total")
    ).filter(
        Sufragante.fecha_registro >= datetime.utcnow() - timedelta(days=dias)
    ).group_by(func.date(Sufragante.fecha_registro)).order_by("fecha").all()

    return [{"fecha": str(row[0]), "total": row[1]} for row in resultados]

@app.get("/export/dashboard/xlsx")
async def export_dashboard_excel(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Exportar datos del dashboard a Excel: resumen KPIs, sufragantes por municipio,
    sufragantes por líder. Solo para superadmin (botón Exportar Excel del Dashboard).
    """
    # Métricas (misma lógica que get_dashboard)
    total_lideres = db.query(func.count(Lider.id)).filter(Lider.activo == True).scalar()
    total_sufragantes = db.query(func.count(Sufragante.id)).scalar()
    verificados = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "verificado").scalar()
    inconsistentes = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "inconsistente").scalar()
    en_revision = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "revision").scalar()
    sin_verificar = db.query(func.count(Sufragante.id)).filter(Sufragante.estado_validacion == "sin_verificar").scalar()
    hoy = datetime.utcnow().date()
    semana_inicio = hoy - timedelta(days=hoy.weekday())
    mes_inicio = hoy.replace(day=1)
    registros_hoy = db.query(func.count(Sufragante.id)).filter(func.date(Sufragante.fecha_registro) == hoy).scalar()
    registros_semana = db.query(func.count(Sufragante.id)).filter(Sufragante.fecha_registro >= semana_inicio).scalar()
    registros_mes = db.query(func.count(Sufragante.id)).filter(Sufragante.fecha_registro >= mes_inicio).scalar()

    # Sufragantes por municipio
    rows_mun = db.query(
        Sufragante.municipio,
        func.count(Sufragante.id).label("total")
    ).group_by(Sufragante.municipio).having(Sufragante.municipio != None).all()

    # Sufragantes por líder (tabla completa)
    rows_lider = db.query(
        Lider.id,
        Lider.nombre,
        Lider.cedula,
        func.count(Sufragante.id).label("total"),
        func.count(case((Sufragante.estado_validacion == "verificado", 1), else_=None)).label("verificados"),
        func.count(case((Sufragante.estado_validacion == "sin_verificar", 1), else_=None)).label("sin_verificar"),
        func.count(case((Sufragante.estado_validacion == "revision", 1), else_=None)).label("en_revision"),
        func.count(case((Sufragante.estado_validacion == "inconsistente", 1), else_=None)).label("inconsistentes"),
    ).outerjoin(Sufragante, Lider.id == Sufragante.lider_id).group_by(Lider.id).all()

    wb = openpyxl.Workbook()

    # Hoja Resumen (KPIs del dashboard)
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.append(["Indicador", "Valor"])
    ws_resumen.append(["Total Líderes", total_lideres or 0])
    ws_resumen.append(["Total Sufragantes", total_sufragantes or 0])
    ws_resumen.append(["Verificados", verificados or 0])
    ws_resumen.append(["Inconsistentes", inconsistentes or 0])
    ws_resumen.append(["En revisión", en_revision or 0])
    ws_resumen.append(["Sin verificar", sin_verificar or 0])
    ws_resumen.append([])
    ws_resumen.append(["Registros Hoy", registros_hoy or 0])
    ws_resumen.append(["Registros Esta Semana", registros_semana or 0])
    ws_resumen.append(["Registros Este Mes", registros_mes or 0])

    # Hoja Sufragantes por Municipio
    ws_mun = wb.create_sheet("Sufragantes por Municipio")
    ws_mun.append(["MUNICIPIO", "TOTAL"])
    for row in rows_mun:
        ws_mun.append([row[0] or "Sin especificar", row[1]])

    # Hoja Sufragantes por Líder (tabla completa)
    ws_lider = wb.create_sheet("Sufragantes por Líder")
    ws_lider.append(["LÍDER", "CÉDULA", "TOTAL", "VERIF.", "SIN VERIF.", "REVISIÓN", "INCONS."])
    for row in rows_lider:
        ws_lider.append([
            row[1] or "",
            row[2] or "",
            row[3] or 0,
            row[4] or 0,
            row[5] or 0,
            row[6] or 0,
            row[7] or 0
        ])

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    buffer.close()

    filename = f"dashboard_innovabigdata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    logger.info(f"Usuario {current_user.username} exportó dashboard a Excel")

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/export/leaders/xlsx")
async def export_leaders_excel(
    leader_ids: Optional[str] = Query(None, description="IDs de líderes separados por coma; si no se envía, se exportan todos"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Exportar registros por líder y totales a Excel. Solo superadmin.
    Si leader_ids no se envía, se exportan todos los líderes activos.
    """
    query_lideres = db.query(Lider).filter(Lider.activo == True).order_by(Lider.nombre)
    if leader_ids and leader_ids.strip():
        try:
            ids = [int(x.strip()) for x in leader_ids.split(",") if x.strip()]
            if ids:
                query_lideres = query_lideres.filter(Lider.id.in_(ids))
        except ValueError:
            raise HTTPException(status_code=400, detail="leader_ids debe ser una lista de números separados por coma")

    lideres = query_lideres.all()
    if not lideres:
        raise HTTPException(status_code=404, detail="No hay líderes para exportar")

    # Resumen por líder (totales)
    wb = openpyxl.Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen por líder"
    ws_resumen.append(["Líder", "Cédula", "Total sufragantes", "Verificados", "Sin verificar", "En revisión", "Inconsistentes"])
    for lid in lideres:
        counts = db.query(
            func.count(Sufragante.id).label("total"),
            func.count(case((Sufragante.estado_validacion == "verificado", 1), else_=None)).label("verificados"),
            func.count(case((Sufragante.estado_validacion == "sin_verificar", 1), else_=None)).label("sin_verificar"),
            func.count(case((Sufragante.estado_validacion == "revision", 1), else_=None)).label("en_revision"),
            func.count(case((Sufragante.estado_validacion == "inconsistente", 1), else_=None)).label("inconsistentes"),
        ).filter(Sufragante.lider_id == lid.id).first()
        ws_resumen.append([
            lid.nombre or "",
            lid.cedula or "",
            counts[0] or 0,
            counts[1] or 0,
            counts[2] or 0,
            counts[3] or 0,
            counts[4] or 0,
        ])

    # Detalle: sufragantes de los líderes seleccionados
    ws_detalle = wb.create_sheet("Sufragantes")
    ws_detalle.append(["Líder", "Nombre", "Cédula", "Edad", "Celular", "Departamento", "Municipio", "Lugar votación", "Mesa", "Estado", "Fecha registro"])
    lid_ids = [l.id for l in lideres]
    sufragantes = db.query(Sufragante).filter(Sufragante.lider_id.in_(lid_ids)).order_by(Sufragante.lider_id, Sufragante.fecha_registro.desc()).all()
    lideres_by_id = {l.id: l for l in lideres}
    for s in sufragantes:
        lid = lideres_by_id.get(s.lider_id)
        lid_nombre = lid.nombre if lid else ""
        ws_detalle.append([
            lid_nombre,
            s.nombre or "",
            s.cedula or "",
            s.edad,
            s.celular or "",
            s.departamento or "",
            s.municipio or "",
            s.lugar_votacion or "",
            s.mesa_votacion or "",
            s.estado_validacion or "",
            s.fecha_registro.strftime("%Y-%m-%d %H:%M") if s.fecha_registro else "",
        ])

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    buffer.close()
    filename = f"lideres_sufragantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    logger.info(f"Usuario {current_user.username} exportó Excel de líderes")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/export/xlsx")
async def export_excel(
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(get_current_user)
):
    """
    Exportar todos los sufragantes (y líderes) a Excel. Permitido para superadmin y operador.
    """
    if current_user.rol not in ("superadmin", "operador"):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sin permiso para exportar")

    sufragantes = db.query(Sufragante).order_by(Sufragante.fecha_registro.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sufragantes"

    headers = [
        "ID", "Nombre", "Cédula", "Edad", "Celular", "Dirección Residencia",
        "Género", "Departamento", "Municipio", "Lugar Votación", "Mesa",
        "Dirección Puesto", "Estado Validación", "Líder ID", "Usuario Registro",
        "Fecha Registro"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row_idx, s in enumerate(sufragantes, 2):
        ws.cell(row=row_idx, column=1, value=s.id)
        ws.cell(row=row_idx, column=2, value=s.nombre or "")
        ws.cell(row=row_idx, column=3, value=s.cedula or "")
        ws.cell(row=row_idx, column=4, value=s.edad)
        ws.cell(row=row_idx, column=5, value=s.celular or "")
        ws.cell(row=row_idx, column=6, value=s.direccion_residencia or "")
        ws.cell(row=row_idx, column=7, value=s.genero or "")
        ws.cell(row=row_idx, column=8, value=s.departamento or "")
        ws.cell(row=row_idx, column=9, value=s.municipio or "")
        ws.cell(row=row_idx, column=10, value=s.lugar_votacion or "")
        ws.cell(row=row_idx, column=11, value=s.mesa_votacion or "")
        ws.cell(row=row_idx, column=12, value=s.direccion_puesto or "")
        ws.cell(row=row_idx, column=13, value=s.estado_validacion or "")
        ws.cell(row=row_idx, column=14, value=s.lider_id)
        ws.cell(row=row_idx, column=15, value=s.usuario_registro or "")
        ws.cell(row=row_idx, column=16, value=s.fecha_registro.isoformat() if s.fecha_registro else "")

    ws2 = wb.create_sheet("Líderes")
    lideres = db.query(Lider).order_by(Lider.nombre).all()
    headers_lideres = [
        "ID", "Nombre", "Cédula", "Edad", "Celular", "Dirección",
        "Género", "Departamento", "Municipio", "Barrio", "Zona Influencia",
        "Tipo Liderazgo", "Usuario Registro", "Fecha Registro", "Activo"
    ]
    for col, header in enumerate(headers_lideres, 1):
        ws2.cell(row=1, column=col, value=header)
    for row_idx, l in enumerate(lideres, 2):
        ws2.cell(row=row_idx, column=1, value=l.id)
        ws2.cell(row=row_idx, column=2, value=l.nombre or "")
        ws2.cell(row=row_idx, column=3, value=l.cedula or "")
        ws2.cell(row=row_idx, column=4, value=l.edad)
        ws2.cell(row=row_idx, column=5, value=l.celular or "")
        ws2.cell(row=row_idx, column=6, value=l.direccion or "")
        ws2.cell(row=row_idx, column=7, value=l.genero or "")
        ws2.cell(row=row_idx, column=8, value=l.departamento or "")
        ws2.cell(row=row_idx, column=9, value=l.municipio or "")
        ws2.cell(row=row_idx, column=10, value=l.barrio or "")
        ws2.cell(row=row_idx, column=11, value=l.zona_influencia or "")
        ws2.cell(row=row_idx, column=12, value=l.tipo_liderazgo or "")
        ws2.cell(row=row_idx, column=13, value=l.usuario_registro or "")
        ws2.cell(row=row_idx, column=14, value=l.fecha_registro.isoformat() if l.fecha_registro else "")
        ws2.cell(row=row_idx, column=15, value="Sí" if l.activo else "No")

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    buffer.close()

    filename = f"exportacion_innovabigdata_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    logger.info(f"Usuario {current_user.username} ({current_user.rol}) exportó datos a Excel")

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/export/incidencias/xlsx")
async def export_incidencias_excel(
    leader_ids: Optional[str] = Query(None, description="IDs de líderes separados por coma; si no se envía, se exportan todos"),
    db: Session = Depends(get_db),
    current_user: TokenData = Depends(require_superadmin)
):
    """
    Exportar incidencias (errores) de cargas masivas por líder a Excel. Solo superadmin.
    Si leader_ids no se envía, se exportan las incidencias de todos los líderes.
    """
    query = db.query(CargaMasivaIncidencia, Lider).join(Lider, Lider.id == CargaMasivaIncidencia.lider_id)
    selected_ids: Optional[List[int]] = None
    if leader_ids and leader_ids.strip():
        try:
            selected_ids = [int(x.strip()) for x in leader_ids.split(",") if x.strip()]
            if selected_ids:
                query = query.filter(CargaMasivaIncidencia.lider_id.in_(selected_ids))
        except ValueError:
            raise HTTPException(status_code=400, detail="leader_ids debe ser una lista de números separados por coma")

    rows = query.order_by(CargaMasivaIncidencia.fecha.desc(), CargaMasivaIncidencia.id.desc()).all()
    if not rows:
        raise HTTPException(status_code=404, detail="No hay incidencias registradas para exportar")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Incidencias"
    ws.append(["Fecha", "Líder", "Cédula líder", "Archivo", "Creados", "Filas archivo", "Fila (Excel)", "Error"])

    fila_re = re.compile(r"^Fila\\s+(\\d+)\\s*:\\s*(.*)$", re.IGNORECASE)
    for inc, lid in rows:
        try:
            errores = json.loads(inc.errores_json or "[]")
        except Exception:
            errores = []
        if not isinstance(errores, list):
            errores = []
        if not errores:
            ws.append([
                inc.fecha.strftime("%Y-%m-%d %H:%M:%S") if inc.fecha else "",
                lid.nombre or "",
                lid.cedula or "",
                inc.archivo or "",
                inc.created or 0,
                inc.total_rows or 0,
                "",
                "",
            ])
            continue
        for e in errores:
            s = str(e or "")
            m = fila_re.match(s)
            fila_excel = int(m.group(1)) if m else None
            msg = (m.group(2) if m else s).strip()
            ws.append([
                inc.fecha.strftime("%Y-%m-%d %H:%M:%S") if inc.fecha else "",
                lid.nombre or "",
                lid.cedula or "",
                inc.archivo or "",
                inc.created or 0,
                inc.total_rows or 0,
                fila_excel,
                msg,
            ])

    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()
    buffer.close()

    def _safe_name(name: str) -> str:
        base = re.sub(r"\\s+", "_", (name or "").strip())
        base = re.sub(r"[^A-Za-z0-9_\\-]+", "", base)
        return base or "lider"

    suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
    if selected_ids and len(selected_ids) == 1:
        lid = db.query(Lider).filter(Lider.id == selected_ids[0]).first()
        filename = f"incidencias_{_safe_name(lid.nombre if lid else 'lider')}_{suffix}.xlsx"
    elif selected_ids:
        filename = f"incidencias_{len(selected_ids)}_lideres_{suffix}.xlsx"
    else:
        filename = f"incidencias_todos_{suffix}.xlsx"

    logger.info(f"Usuario {current_user.username} exportó incidencias a Excel")
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'}
    )

# =====================
# SALUD DE LA APLICACIÓN
# =====================

@app.get("/health")
async def health_check():
    """
    Endpoint de verificación de salud
    """
    try:
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        return {"status": "unhealthy", "error": str(e)}

@app.get("/")
async def root():
    """
    Endpoint raíz
    """
    return {
        "message": "API Sistema de Registro de Líderes y Sufragantes",
        "version": "1.0.0",
        "status": "operational"
    }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
